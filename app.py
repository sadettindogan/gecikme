import streamlit as st
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook, Workbook
import os
import time
import tempfile
import math
import zipfile
import io
import re

# --- SAYFA AYARLARI ---
st.set_page_config(page_title="Gecikme Zammı Otomasyonu", page_icon="📄")

# --- TARİH FORMATI ---
def tarih_str(t):
    return t.strftime("%d.%m.%Y") if hasattr(t, "strftime") else str(t)

# --- TUTAR STR ---
def miktar_ham_str(deger):
    if isinstance(deger, float):
        return str(deger).replace(".", ",")
    if isinstance(deger, int):
        return str(deger)
    return str(deger).strip()

# --- VALİDASYON ---
def miktar_dogrula(satirlar):
    hatalar = []
    for idx, (a, b, c) in enumerate(satirlar):
        satir_no = idx + 1
        ham   = str(a).strip()
        deger = miktar_ham_str(a)
        if "." in deger:
            hatalar.append((satir_no, ham, "Nokta (.) içeriyor — ondalık ayracı virgül (,) olmalıdır."))
            continue
        if "," in deger:
            parcalar = deger.split(",")
            if len(parcalar) > 2:
                hatalar.append((satir_no, ham, "Birden fazla virgül içeriyor."))
                continue
            elif len(parcalar[1]) > 2:
                hatalar.append((satir_no, ham, f"Ondalık kısmı {len(parcalar[1])} hane — en fazla 2 hane olabilir."))
                continue
        temiz = deger.replace(",", "")
        if not temiz.isdigit():
            hatalar.append((satir_no, ham, "Geçersiz karakter (yalnızca rakam ve virgül kabul edilir)."))
    return hatalar

# --- PARSE ---
TARIH_PATTERN = re.compile(r'\b(\d{1,2})[./](\d{1,2})[./](\d{4})\b')

def tarih_normalize(s):
    m = TARIH_PATTERN.match(s.strip())
    if m:
        gun, ay, yil = m.group(1), m.group(2), m.group(3)
        return f"{gun.zfill(2)}.{ay.zfill(2)}.{yil}"
    return s.strip()

def parse_yapistirilmis_metin(metin):
    satirlar = []
    hatali_satirlar = []
    lines = metin.strip().splitlines()
    for i, line in enumerate(lines):
        line = line.strip()
        if not line:
            continue
        hucreler = line.split("\t")
        if len(hucreler) >= 3:
            tutar = hucreler[0].strip()
            vade  = tarih_normalize(hucreler[1].strip())
            odeme = tarih_normalize(hucreler[2].strip())
            if tutar and vade and odeme:
                satirlar.append((tutar, vade, odeme))
                continue
        tarihler = TARIH_PATTERN.findall(line)
        if len(tarihler) >= 2:
            ilk_eslesme = TARIH_PATTERN.search(line)
            tutar = line[:ilk_eslesme.start()].strip().rstrip("\t ,")
            gun1, ay1, yil1 = tarihler[0]
            gun2, ay2, yil2 = tarihler[1]
            vade  = f"{gun1.zfill(2)}.{ay1.zfill(2)}.{yil1}"
            odeme = f"{gun2.zfill(2)}.{ay2.zfill(2)}.{yil2}"
            if tutar and vade and odeme:
                satirlar.append((tutar, vade, odeme))
                continue
            else:
                hatali_satirlar.append((i + 1, line, "Tutar ayrıştırılamadı."))
                continue
        hatali_satirlar.append((i + 1, line, "3 sütun bulunamadı. Sekme ayracı eksik olabilir."))
    return satirlar, hatali_satirlar

def wb_olustur(satirlar):
    wb = Workbook()
    sheet = wb.active
    for idx, (a, b, c) in enumerate(satirlar):
        sheet[f"A{idx+1}"] = a
        sheet[f"B{idx+1}"] = b
        sheet[f"C{idx+1}"] = c
    return wb

# ============================================================
# SESSION STATE
# ============================================================
for key, default in [
    ("zip_bytes", None),
    ("sonuc_satirlar", []),
    ("satirlar_cache", []),
]:
    if key not in st.session_state:
        st.session_state[key] = default

# ============================================================
# BAŞLIK
# ============================================================
st.title("📄 Gecikme Zammı Rapor Portalı")
st.markdown(
    "Satır sınırı yoktur. Başlık satırı olmadan **Tutar(*) / Vade Tarihi / Ödeme Tarihi** "
    "sütunlarını Excel'den kopyalayıp aşağıdaki kutuya yapıştırın **(Ctrl+V)**. "
    "Sonucu kopyalayıp aynı yere yapıştırabilirsiniz. "
    "GİB Sitesi çıktılarına **İndir** butonu ile ulaşılabilir.\n\n"
    "*(\*) Tutar verisi şu kurala uymazsa hata verecektir: "
    "Nokta içermemeli, ondalık varsa virgül ile ayrılmalı, virgülden sonra en fazla iki hane olmalı.*"
)

# ============================================================
# GİRİŞ
# ============================================================
yapistir_metni = st.text_area(
    "",
    height=160,
    placeholder="1500\t01.01.2023\t15.06.2023\n2300\t15.03.2023\t20.09.2023",
    label_visibility="collapsed",
    key="ta_yapistir",
)

enter_tiklandi = st.button("↵ Enter (Veriyi Kontrol Ederek Hata Denetiminden Geçirir.)")

# ============================================================
# ENTER — PARSE + VALİDASYON
# ============================================================
if enter_tiklandi:
    st.session_state.sonuc_satirlar = []
    st.session_state.zip_bytes      = None
    st.session_state.satirlar_cache = []

    if not yapistir_metni.strip():
        st.warning("⚠️ Lütfen veri yapıştırın.")
    else:
        satirlar, parse_hatalari = parse_yapistirilmis_metin(yapistir_metni)

        if parse_hatalari:
            st.error(f"❌ {len(parse_hatalari)} satırda format hatası:")
            st.table([{"Satır": s, "İçerik": ic, "Hata": h} for s, ic, h in parse_hatalari])
            st.stop()

        if not satirlar:
            st.error("Geçerli satır bulunamadı.")
            st.stop()

        hatalar_val = miktar_dogrula(satirlar)
        if hatalar_val:
            st.error(
                f"❌ {len(hatalar_val)} satırda tutar hatası bulundu. "
                "GİB sistemi veriyi kabul etmemektedir."
            )
            st.table([{"Satır": s, "Girilen Değer": d, "Hata": h} for s, d, h in hatalar_val])
            st.stop()

        # Hata yoksa cache'e al ve başlat butonunu göster
        st.session_state.satirlar_cache = satirlar
        st.success(f"✅ {len(satirlar)} satır geçerli.")

# ============================================================
# BAŞLAT — sadece cache doluysa göster
# ============================================================
if st.session_state.satirlar_cache:
    if st.button("🚀 Başlat", type="primary"):
        satirlar       = st.session_state.satirlar_cache
        tmp_dir        = tempfile.mkdtemp()
        wb_orijinal    = wb_olustur(satirlar)
        sheet_orijinal = wb_orijinal.active
        MAX_GRUP       = 25
        grup_sayisi    = math.ceil(len(satirlar) / MAX_GRUP)
        durum_alani    = st.empty()
        durum_alani.markdown("🟡 **%0 — İşleniyor...**")
        sonuclar = {}

        try:
            with sync_playwright() as p:
                browser = p.chromium.launch(
                    headless=True,
                    executable_path="/usr/bin/chromium",
                    args=["--no-sandbox", "--disable-dev-shm-usage"]
                )
                context = browser.new_context(accept_downloads=True)
                page    = context.new_page()

                page.goto("https://dijital.gib.gov.tr/hesaplamalar/GecikmeZamVeFaizHesaplama")
                page.wait_for_load_state("networkidle")

                def satir_sayisi():
                    return len(page.query_selector_all("input[id^='odenecekMiktar']"))

                def yeni_satir_ekle():
                    once = satir_sayisi()
                    for deneme in range(5):
                        btn = page.query_selector("button[aria-label='add']")
                        if not btn:
                            time.sleep(0.5)
                            continue
                        btn.scroll_into_view_if_needed()
                        btn.click()
                        deadline = time.time() + 5
                        while time.time() < deadline:
                            if satir_sayisi() > once:
                                return True
                            time.sleep(0.2)
                    return False

                def formu_sifirla():
                    """Grubu bitirince TEMİZLE butonuna bas, form sıfırlansın."""
                    page.click("#reset")
                    page.wait_for_selector("#odenecekMiktar1", timeout=10000)

                def dropdown_sec(form_index):
                    dropdown_id = f"gecikmeTipi{form_index}"
                    mevcut = page.query_selector(f"input[name='{dropdown_id}']")
                    if mevcut:
                        mevcut_deger = mevcut.get_attribute("value") or ""
                        if "Gecikme Zammı" in mevcut_deger or "gecikmeZammi" in mevcut_deger.lower():
                            return
                    dropdown_div = page.wait_for_selector(f"#{dropdown_id}", timeout=5000)
                    dropdown_div.scroll_into_view_if_needed()
                    dropdown_div.click()
                    time.sleep(0.3)
                    page.wait_for_selector("ul[role='listbox']", timeout=5000)
                    gecikme_li = page.query_selector("li[data-value='Gecikme Zammı']") or \
                                 page.query_selector("li:has-text('Gecikme Zammı')")
                    if gecikme_li:
                        gecikme_li.click()
                    else:
                        page.keyboard.press("Escape")

                def satir_doldur(miktar, vade, odeme, son_mu):
                    form_index = satir_sayisi()
                    dropdown_sec(form_index)
                    inp_miktar = page.wait_for_selector(f"#odenecekMiktar{form_index}", timeout=10000)
                    inp_miktar.click()
                    inp_miktar.fill(miktar_ham_str(miktar))
                    inp_vade = page.wait_for_selector(f"#vadeTarihi{form_index}", timeout=10000)
                    inp_vade.click()
                    inp_vade.fill(tarih_str(vade))
                    page.keyboard.press("Escape")
                    inp_odeme = page.wait_for_selector(f"#odemeTarihi{form_index}", timeout=10000)
                    inp_odeme.click()
                    inp_odeme.fill(tarih_str(odeme))
                    page.keyboard.press("Escape")
                    if not son_mu:
                        if not yeni_satir_ekle():
                            st.warning("Yeni satır eklenemedi, işlem durdu.")
                            return False
                    return True

                for grup_no in range(grup_sayisi):
                    baslangic = grup_no * MAX_GRUP
                    bitis     = min(baslangic + MAX_GRUP, len(satirlar))
                    grup      = satirlar[baslangic:bitis]
                    etiket    = f"{baslangic + 1}-{bitis}"
                    yuzde     = int((grup_no / grup_sayisi) * 100)
                    durum_alani.markdown(f"🟡 **%{yuzde} — İşleniyor...**")

                    for idx, (miktar, vade, odeme) in enumerate(grup):
                        son_mu = (idx == len(grup) - 1)
                        ok = satir_doldur(miktar, vade, odeme, son_mu)
                        if not ok:
                            break

                    page.wait_for_selector("#submit:enabled", timeout=10000)
                    page.click("#submit")
                    page.wait_for_selector("#exportPdfButton:enabled", timeout=30000)

                    pdf_yolu = os.path.join(tmp_dir, f"xvb_{etiket}.pdf")
                    with page.expect_download() as dl_info:
                        page.click("#exportPdfButton")
                    dl_info.value.save_as(pdf_yolu)

                    excel_yolu = os.path.join(tmp_dir, f"xvb_{etiket}.xlsx")
                    with page.expect_download() as xl_info:
                        page.get_by_text("Excel'e Aktar").click()
                    xl_info.value.save_as(excel_yolu)

                    # GİB Excel G sütunu
                    wb_gib    = load_workbook(excel_yolu, data_only=True)
                    sheet_gib = wb_gib.active
                    for i in range(len(grup)):
                        g_degeri = sheet_gib[f"G{3 + i}"].value
                        if g_degeri is not None:
                            try:
                                g_str = f"{float(str(g_degeri).replace(',', '.')):.2f}".replace(".", ",")
                            except (ValueError, TypeError):
                                g_str = str(g_degeri)
                        else:
                            g_str = None
                        hucre = sheet_orijinal.cell(row=baslangic + 1 + i, column=4, value=g_str)
                        hucre.number_format = "@"
                        a_val, b_val, c_val = grup[i]
                        st.session_state.sonuc_satirlar.append((
                            miktar_ham_str(a_val),
                            tarih_str(b_val),
                            tarih_str(c_val),
                            g_str or "",
                        ))

                    with open(pdf_yolu, "rb") as f:
                        sonuclar[f"xvb_{etiket}.pdf"] = f.read()
                    with open(excel_yolu, "rb") as f:
                        sonuclar[f"xvb_{etiket}.xlsx"] = f.read()
                    try:
                        os.remove(pdf_yolu)
                        os.remove(excel_yolu)
                    except Exception:
                        pass

                    yuzde_bitti = int(((grup_no + 1) / grup_sayisi) * 100)
                    durum_alani.markdown(f"🟡 **%{yuzde_bitti} — İşleniyor...**")

                    # Son grup değilse formu sıfırla
                    if grup_no < grup_sayisi - 1:
                        formu_sifirla()

                browser.close()

            sonuc_buffer = io.BytesIO()
            wb_orijinal.save(sonuc_buffer)
            sonuclar["sonuc_dosyasi.xlsx"] = sonuc_buffer.getvalue()

            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                for dosya_adi, icerik in sonuclar.items():
                    zf.writestr(dosya_adi, icerik)
            st.session_state.zip_bytes = zip_buffer.getvalue()

            durum_alani.markdown("🟢 **%100 — Tamamlandı!**")

        except Exception as e:
            st.error(f"❌ Bir hata oluştu: {str(e)}")

# ============================================================
# SONUÇ
# ============================================================
if st.session_state.sonuc_satirlar:
    st.markdown("---")
    tsv = "\n".join("\t".join(row) for row in st.session_state.sonuc_satirlar)
    st.code(tsv, language=None)

    tsv_js = tsv.replace("\\", "\\\\").replace("`", "\\`")
    st.components.v1.html(
        f"""
        <button onclick="
            navigator.clipboard.writeText(`{tsv_js}`).then(function() {{
                this.innerText = '\u2705 Kopyaland\u0131!';
                setTimeout(() => this.innerText = '\U0001f4cb Sonucu Kopyala', 2000);
            }}.bind(this)).catch(function() {{
                this.innerText = '\u274c Kopyalanamad\u0131';
                setTimeout(() => this.innerText = '\U0001f4cb Sonucu Kopyala', 2000);
            }}.bind(this));
        " style="
            padding:8px 18px;
            font-size:14px;
            cursor:pointer;
            border:1px solid #ccc;
            border-radius:6px;
            background:#f0f2f6;
            color:#333;
        ">\U0001f4cb Sonucu Kopyala</button>
        """,
        height=48,
    )

if st.session_state.zip_bytes:
    st.download_button(
        label="📦 İndir (ZIP)",
        data=st.session_state.zip_bytes,
        file_name="xvb_raporlar.zip",
        mime="application/zip"
    )
